import io
import csv
import json
import openpyxl

def parse_csv_data(file_bytes: bytes) -> dict:
    text = file_bytes.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return {"file_type": "CSV", "rows_count": 0, "columns": [], "data": []}
    
    headers = rows[0]
    data_rows = rows[1:]
    
    # Analyze data types
    columns = analyze_columns(headers, data_rows)
    
    # Format first 1000 rows as list of dicts
    data = []
    for r in data_rows[:1000]:
        row_dict = {}
        for idx, col in enumerate(headers):
            val = r[idx] if idx < len(r) else ""
            row_dict[col] = val
        data.append(row_dict)
        
    return {
        "file_type": "CSV",
        "rows_count": len(data_rows),
        "columns": columns,
        "data": data
    }

def parse_tsv_data(file_bytes: bytes) -> dict:
    text = file_bytes.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text), delimiter="\t")
    rows = list(reader)
    if not rows:
        return {"file_type": "TSV", "rows_count": 0, "columns": [], "data": []}
    
    headers = rows[0]
    data_rows = rows[1:]
    
    columns = analyze_columns(headers, data_rows)
    
    data = []
    for r in data_rows[:1000]:
        row_dict = {}
        for idx, col in enumerate(headers):
            val = r[idx] if idx < len(r) else ""
            row_dict[col] = val
        data.append(row_dict)
        
    return {
        "file_type": "TSV",
        "rows_count": len(data_rows),
        "columns": columns,
        "data": data
    }

def parse_json_data(file_bytes: bytes) -> dict:
    text = file_bytes.decode("utf-8", errors="ignore")
    parsed = json.loads(text)
    
    # Handle both: list of dicts or dict of lists
    if isinstance(parsed, dict):
        # If it's a dict with list values, try to reconstruct tabular structure
        for key, val in parsed.items():
            if isinstance(val, list) and len(val) > 0 and isinstance(val[0], dict):
                parsed = val
                break
                
    if not isinstance(parsed, list):
        # Fallback to wrapping single item
        parsed = [parsed]
        
    if not parsed:
        return {"file_type": "JSON", "rows_count": 0, "columns": [], "data": []}
        
    # Get all keys across all records to build headers
    headers_set = set()
    for row in parsed:
        if isinstance(row, dict):
            headers_set.update(row.keys())
    headers = sorted(list(headers_set))
    
    data_rows = []
    for row in parsed:
        if isinstance(row, dict):
            data_rows.append([str(row.get(h, "")) for h in headers])
            
    columns = analyze_columns(headers, data_rows)
    
    # Format data
    data = []
    for row in parsed[:1000]:
        if isinstance(row, dict):
            data.append({h: str(row.get(h, "")) for h in headers})
            
    return {
        "file_type": "JSON",
        "rows_count": len(parsed),
        "columns": columns,
        "data": data
    }

def parse_excel_data(file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet_names = wb.sheetnames
    
    # Default to first sheet
    ws = wb[sheet_names[0]]
    rows = list(ws.iter_rows(values_only=True))
    
    if not rows:
        return {"file_type": "Excel", "rows_count": 0, "columns": [], "data": [], "sheets": sheet_names}
        
    headers = [str(cell) if cell is not None else f"Column_{i}" for i, cell in enumerate(rows[0])]
    data_rows = []
    for r in rows[1:]:
        data_rows.append([str(cell) if cell is not None else "" for cell in r])
        
    columns = analyze_columns(headers, data_rows)
    
    data = []
    for r in data_rows[:1000]:
        row_dict = {}
        for idx, col in enumerate(headers):
            val = r[idx] if idx < len(r) else ""
            row_dict[col] = val
        data.append(row_dict)
        
    return {
        "file_type": "Excel",
        "rows_count": len(data_rows),
        "columns": columns,
        "data": data,
        "sheets": sheet_names
    }

def analyze_columns(headers: list, rows: list) -> list:
    """
    Deduce column data types by looking at content
    """
    columns = []
    for idx, col_name in enumerate(headers):
        # sample values
        sample_vals = []
        for r in rows[:100]:
            if idx < len(r) and r[idx].strip() != "":
                sample_vals.append(r[idx].strip())
                
        # Classify as numeric or text
        is_numeric = True
        if not sample_vals:
            is_numeric = False
        else:
            for val in sample_vals:
                try:
                    float(val)
                except ValueError:
                    is_numeric = False
                    break
                    
        columns.append({
            "name": col_name,
            "type": "numeric" if is_numeric else "text"
        })
    return columns

def parse_uploaded_data(file_bytes: bytes, filename: str) -> dict:
    fn = filename.lower()
    if fn.endswith(".csv"):
        return parse_csv_data(file_bytes)
    elif fn.endswith(".tsv"):
        return parse_tsv_data(file_bytes)
    elif fn.endswith(".json"):
        return parse_json_data(file_bytes)
    elif fn.endswith(".xlsx") or fn.endswith(".xls"):
        return parse_excel_data(file_bytes)
    else:
        raise ValueError(f"Unsupported file type: {filename}")
